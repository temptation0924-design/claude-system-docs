#!/usr/bin/env python3
"""남양리 2055-3번지 물건현황표 엑셀 생성"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── 공통 스타일 ──
HEADER_FILL = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")  # 빨간색
HEADER_FONT = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="맑은 고딕", size=14, bold=True)
SUBTITLE_FONT = Font(name="맑은 고딕", size=11, bold=True)
NORMAL_FONT = Font(name="맑은 고딕", size=11)
ACCENT_FONT = Font(name="맑은 고딕", size=12, bold=True, color="C0392B")
WON_FORMAT = '#,##0'
PCT_FORMAT = '0.0%'

BLUE_FILL = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
BLUE_FONT = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
LIGHT_GRAY = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

CENTER = Alignment(horizontal='center', vertical='center')
RIGHT = Alignment(horizontal='right', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')


def apply_border(ws, row_range, col_range):
    for r in row_range:
        for c in col_range:
            ws.cell(row=r, column=c).border = thin_border


# ══════════════════════════════════════════
# Sheet 1: 투자분석표
# ══════════════════════════════════════════
ws1 = wb.active
ws1.title = "투자분석표"
ws1.sheet_properties.tabColor = "C0392B"

# 열 너비
ws1.column_dimensions['A'].width = 4
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 22
ws1.column_dimensions['D'].width = 22

# 타이틀
ws1.merge_cells('B2:D2')
ws1['B2'] = "남양리 2055-3번지 투자분석표"
ws1['B2'].font = TITLE_FONT
ws1['B2'].alignment = CENTER

# 헤더
headers = ['항목', '금액', '비고']
for i, h in enumerate(headers, start=2):
    cell = ws1.cell(row=4, column=i, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER

# 데이터
data = [
    ('매매가액', 1400000000, ''),
    ('대출금액', 1070000000, ''),
    ('금리', 0.052, '5.8% (비고)'),
    ('보증금', 146000000, ''),
    ('전체건물수익', 10000000, '월'),
    ('한달이자', 4636667, ''),
    ('각종공과금', 700000, ''),
    ('비용공제수익', 4663333, ''),
    ('인수비용', 184000000, '₩55,960,000 (비고)'),
    ('수익률', 0.3041, ''),
]

for i, (item, amount, note) in enumerate(data, start=5):
    ws1.cell(row=i, column=2, value=item).font = SUBTITLE_FONT
    ws1.cell(row=i, column=2).alignment = CENTER

    cell_amt = ws1.cell(row=i, column=3, value=amount)
    if isinstance(amount, float) and amount < 1:
        cell_amt.number_format = '0.0%'
        if item == '수익률':
            cell_amt.font = ACCENT_FONT
        else:
            cell_amt.font = NORMAL_FONT
    else:
        cell_amt.number_format = '₩#,##0'
        cell_amt.font = NORMAL_FONT
    cell_amt.alignment = RIGHT

    ws1.cell(row=i, column=4, value=note).font = NORMAL_FONT
    ws1.cell(row=i, column=4).alignment = CENTER

    # 짝수 행 배경
    if i % 2 == 0:
        for c in range(2, 5):
            ws1.cell(row=i, column=c).fill = LIGHT_GRAY

# 테두리
apply_border(ws1, range(4, 15), range(2, 5))


# ══════════════════════════════════════════
# Sheet 2: 호실현황표
# ══════════════════════════════════════════
ws2 = wb.create_sheet("호실현황표")
ws2.sheet_properties.tabColor = "3498DB"

# 열 너비
ws2.column_dimensions['A'].width = 4
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 10
ws2.column_dimensions['D'].width = 10
ws2.column_dimensions['E'].width = 20
ws2.column_dimensions['F'].width = 20

# 타이틀
ws2.merge_cells('B2:F2')
ws2['B2'] = "경기도 화성시 남양리 2055-3번지 호실현황"
ws2['B2'].font = TITLE_FONT
ws2['B2'].alignment = CENTER

# 헤더
headers2 = ['호실', '층수', '방개수', '월세+관리비', '보증금']
for i, h in enumerate(headers2, start=2):
    cell = ws2.cell(row=4, column=i, value=h)
    cell.font = BLUE_FONT
    cell.fill = BLUE_FILL
    cell.alignment = CENTER

# 데이터
rooms = [
    ('상가', 1, 1, 1000000, 10000000),
    ('201', 2, 4, 2000000, 4000000),
    ('202', 2, 4, 2000000, 4000000),
    ('301', 3, 4, 2000000, 4000000),
    ('302', 3, 4, 2000000, 4000000),
    ('401', 4, 5, 1000000, 120000000),
]

for i, (room, floor, rooms_cnt, rent, deposit) in enumerate(rooms, start=5):
    ws2.cell(row=i, column=2, value=room).font = SUBTITLE_FONT
    ws2.cell(row=i, column=2).alignment = CENTER
    ws2.cell(row=i, column=3, value=floor).font = NORMAL_FONT
    ws2.cell(row=i, column=3).alignment = CENTER
    ws2.cell(row=i, column=4, value=rooms_cnt).font = NORMAL_FONT
    ws2.cell(row=i, column=4).alignment = CENTER

    cell_rent = ws2.cell(row=i, column=5, value=rent)
    cell_rent.number_format = '₩#,##0'
    cell_rent.font = NORMAL_FONT
    cell_rent.alignment = RIGHT

    cell_dep = ws2.cell(row=i, column=6, value=deposit)
    cell_dep.number_format = '₩#,##0'
    cell_dep.font = NORMAL_FONT
    cell_dep.alignment = RIGHT

    if i % 2 == 0:
        for c in range(2, 7):
            ws2.cell(row=i, column=c).fill = LIGHT_GRAY

# 합계 행
row_total = 11
ws2.cell(row=row_total, column=2, value='합계').font = SUBTITLE_FONT
ws2.cell(row=row_total, column=2).alignment = CENTER
ws2.cell(row=row_total, column=2).fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")

cell_total_rent = ws2.cell(row=row_total, column=5, value=10000000)
cell_total_rent.number_format = '₩#,##0'
cell_total_rent.font = Font(name="맑은 고딕", size=11, bold=True)
cell_total_rent.alignment = RIGHT
cell_total_rent.fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")

cell_total_dep = ws2.cell(row=row_total, column=6, value=146000000)
cell_total_dep.number_format = '₩#,##0'
cell_total_dep.font = Font(name="맑은 고딕", size=11, bold=True)
cell_total_dep.alignment = RIGHT
cell_total_dep.fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")

for c in range(3, 5):
    ws2.cell(row=row_total, column=c).fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")

# 테두리
apply_border(ws2, range(4, 12), range(2, 7))

# ── 저장 ──
output_path = "/Users/ihyeon-u/Downloads/남양리2055-3_물건현황표_v1.xlsx"
wb.save(output_path)
print(f"✅ 저장 완료: {output_path}")
